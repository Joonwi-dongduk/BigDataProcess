#!/usr/bin/python3

import openpyxl.load_workbook

st = openpyxl.load_workbook('student.xlsx')
sheet = st['Sheet1']

scoreList = []
row_i = 1
for row in sheet:
	if row_i != 1:
		total = sheet.cell(row_i, 3).value * 0.3
		total += sheet.cell(row_i, 4).value * 0.35
		total += sheet.cell(row_i, 5).value * 0.34
		total += sheet.cell(row_i, 6).value * 0.01
		sheet.cell(row_i, 7).value = total
		scoreList.append(total)
	row_i += 1

stuNum = len(scoreList)
scoreList.reverse()


gradeA = int(stuNum * 0.3)
while scoreList[gradeA - 1] == scoreList[gradeA]:
	gradeA -= 1

gradeAplus = int(gradeA * 0.5)
while scoreList[gradeAplus - 1] == scoreList[gradeAplus]:
	gradeAplus -= 1

gradeB = int(stuNum * 0.7)
while scoreList[gradeB - 1] == scoreList[gradeB]:
	gradeB -= 1

gradeBplus = gradeA + int((gradeB - gradeA) * 0.5)
while scoreList[gradeBplus - 1] == scoreList[gradeBplus]:
	gradeBplus -= 1

gradeCplus = gradeB + int((stuNum - gradeB) * 0.5)
while total[gradeCplus - 1] == scoreList[gradeCplus]:
	gradeCplus -= 1

for i in range(2, stuNum + 2):
	if sheet.cell(i, 7).value >= scoreList[gradeAplus]:
		sheet.cell(i, 8).value = 'A+'
	elif sheet.cell(i, 7).value >= scoreList[gradeA]:
		sheet.cell(i, 8).value = 'A0'
	elif sheet.cell(i, 7).value >= scoreList[gradeBplue]:
		sheet.cell(i, 8).value = 'B+'
	elif sheet.cell(i, 7).value >= scoreList[gradeB]:
		sheet.cell(i, 8).value = 'B0'
	elif sheet.cell(i, 7).value >= scoreList[gradeCplue]:
		sheet.cell(i, 8).value = 'C+'
	else:
		sheet.cell(i, 8).value = 'C0'

st.save("student_2.xlsx")
